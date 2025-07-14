import openai
import concurrent.futures
import time
import openpyxl
import os
from openpyxl.styles import Font
from datetime import datetime

NUMBER_OF_CLIENTS = 2
EXCEL_PATH = r"src/experiments/multiple_request_benchmark.xlsx"


client = openai.OpenAI(
    base_url="http://192.168.88.130:8080",
    api_key="sk-no-key-required"
)

# TODO : have to clean the code and write enums for below list
COLUMNS = [
    ("session_id", "int"),
    ("BT", "datetime"),
    ("FT", "datetime"),
    ("LT", "datetime"),
    ("prompt_tokens", "int"),
    ("prompt_ms", "float"),
    ("prompt_per_token_ms", "float"),
    ("prompt_per_second", "float"),
    ("predicted_ms", "float"),
    ("predicted_per_token_ms", "float"),
    ("predicted_per_second", "float"),
    ("TTFT", "float"),
    ("TGT", "float"),
    ("PP", "float"),
    ("error", "str"),
    ("content_sample", "str")
]


def init_excel_file():
    directory = os.path.dirname(EXCEL_PATH)
    if not os.path.exists(directory):
        os.makedirs(directory)

    if os.path.exists(EXCEL_PATH):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active

        if sheet.max_row == 0 or sheet.cell(row=1, column=1).value != COLUMNS[0][0]:
            for col_idx, (header, _) in enumerate(COLUMNS, start=1):
                sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
            workbook.save(EXCEL_PATH)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Benchmark Results"

        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)

        workbook.save(EXCEL_PATH)

    return workbook, sheet


def save_to_excel(metrics, workbook, sheet):
    row_data = []
    for col_name, col_type in COLUMNS:
        value = metrics.get(col_name)

        if col_type == "datetime" and value is not None:
            value = datetime.fromtimestamp(value)

        row_data.append(value)


    sheet.append(row_data)

    for col_idx, (_, col_type) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=sheet.max_row, column=col_idx)

        if col_type == "datetime":
            #cell.number_format = 'YYYY-MM-DD HH:MM:SS.000'
            cell.number_format = 'HH:MM:SS.000'
        elif col_type == "float":
            cell.number_format = '0.000'
        elif col_type == "int":
            cell.number_format = '0'

    workbook.save(EXCEL_PATH)


def send_request(session_id, workbook, sheet):
    metrics = {
        'session_id': session_id,
        'BT': time.time(),
        'FT': None,
        'LT': None,
        'prompt_tokens': None,
        'prompt_ms': None,
        'prompt_per_token_ms': None,
        'prompt_per_second': None,
        'predicted_ms': None,
        'predicted_per_token_ms': None,
        'predicted_per_second': None,
        'content': "",
        'error': None
    }

    try:
        stream = client.chat.completions.create(
            model="deepseek-ai/DeepSeek-R1-Distill-Qwen-1.5B",
            messages=[{"role": "system",
                       #"content": "XXXXXXXXXXXXXXX"}],
                       "content": "give me a 1000 word essay"}],
                       #"content": "Translate the following sentence into French: The cat sat on the mat."}],
                       #"content": "A man is getting ready to go for a run. He puts on his... : Options: A) swimming goggles. B) running shoes. C) winter coat. D) chef’s hat. <think> </think>"}],
            # correct answer is B
            stream=True,
            max_tokens=200,
            stream_options={"include_usage": True}
        )

        for chunk in stream:
            if metrics['FT'] is None and chunk.choices[0].delta.content:
                metrics['FT'] = time.time()

            if chunk.choices[0].delta.content:
                metrics['content'] += chunk.choices[0].delta.content
            else:
                if hasattr(chunk, 'usage') and hasattr(chunk, 'model_extra'):
                    usage_data = chunk.usage
                    timings_data = chunk.model_extra.get('timings', {}) if hasattr(chunk, 'model_extra') else {}

                    metrics.update({
                        'prompt_tokens': usage_data.prompt_tokens if hasattr(usage_data, 'prompt_tokens') else None,
                        'prompt_ms': timings_data.get('prompt_ms'),
                        'prompt_per_token_ms': timings_data.get('prompt_per_token_ms'),
                        'prompt_per_second': timings_data.get('prompt_per_second'),
                        'predicted_ms': timings_data.get('predicted_ms'),
                        'predicted_per_token_ms': timings_data.get('predicted_per_token_ms'),
                        'predicted_per_second': timings_data.get('predicted_per_second')
                    })

            if chunk.choices[0].finish_reason:
                metrics['LT'] = time.time()
                break

        if metrics['FT'] and metrics['BT']:
            metrics['TTFT'] = (metrics['FT'] - metrics['BT']) * 1000
        if metrics['LT'] and metrics['FT']:
            metrics['TGT'] = (metrics['LT'] - metrics['FT']) * 1000
        if metrics.get('TTFT') and metrics['prompt_tokens']:
            metrics['PP'] = (metrics['TTFT'] / metrics['prompt_tokens'])

    except Exception as e:
        metrics['error'] = str(e)

    metrics['content_sample'] = (metrics['content'][:100] + "...") if metrics['content'] else None

    save_to_excel(metrics, workbook, sheet)
    return metrics


def run_concurrent_sessions():
    workbook, sheet = init_excel_file()

    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=NUMBER_OF_CLIENTS) as executor:
            futures = {executor.submit(send_request, i, workbook, sheet): i
                       for i in range(NUMBER_OF_CLIENTS)}

            for future in concurrent.futures.as_completed(futures):
                session_id = futures[future]
                try:
                    metrics = future.result()
                    print(f"Session {session_id} completed successfully")
                except Exception as e:
                    print(f"Session {session_id} failed with error: {str(e)}")
    finally:
        workbook.close()


if __name__ == "__main__":
    print(f"Starting {NUMBER_OF_CLIENTS} concurrent client sessions...")
    run_concurrent_sessions()
    print(f"\nResults saved to: {EXCEL_PATH}")
