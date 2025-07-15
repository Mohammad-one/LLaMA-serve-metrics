import random
import openai
import concurrent.futures
import time
import openpyxl
import os
from openpyxl.styles import Font
from datetime import datetime
from src.experiments.models.enums import BenchmarkColumns, ContentType
from transformers import AutoTokenizer

CLIENT_COUNTS = [10, 50, 100]
PROMPT_LENGTHS = [512, 1024, 2048]
CONTENT_TYPE = ContentType.RENDER
MAX_TOKENS = 1
BASE_DIR = r"C:\Users\ASUS\Desktop\Pars\HardwareAware\src\experiments\data\tamrin_tets"

client = openai.OpenAI(
    base_url="http://192.168.70.137:8080",
    api_key="sk-no-key-required"
)

tokenizer = AutoTokenizer.from_pretrained("deepseek-ai/DeepSeek-R1-Distill-Qwen-1.5B")
COLUMNS = BenchmarkColumns.get_all_columns()


def init_excel_file(client_count, prompt_length):
    excel_filename = f"{client_count}_{prompt_length}_{MAX_TOKENS}_{CONTENT_TYPE.value.lower()}.xlsx"
    excel_path = os.path.join(BASE_DIR, excel_filename)

    directory = os.path.dirname(excel_path)
    if not os.path.exists(directory):
        os.makedirs(directory)

    if os.path.exists(excel_path):
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        if sheet.max_row == 0 or sheet.cell(row=1, column=1).value != COLUMNS[0][0]:
            for col_idx, (header, _) in enumerate(COLUMNS, start=1):
                sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
            workbook.save(excel_path)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Benchmark Results"

        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            sheet.cell(row=1, column=col_idx, value=header).font = Font(bold=True)

        workbook.save(excel_path)

    return workbook, sheet, excel_path


def save_to_excel(metrics, workbook, sheet, excel_path):
    row_data = []
    for col_name, col_type in COLUMNS:
        value = metrics.get(col_name)

        if col_type == "datetime" and value is not None:
            value = datetime.fromtimestamp(value)

        row_data.append(value)

    sheet.append(row_data)

    for col_idx, (_, col_type) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=sheet.max_row, column=col_idx)

        if col_type == "datetime":
            cell.number_format = 'HH:MM:SS.000'
        elif col_type == "float":
            cell.number_format = '0.000'
        elif col_type == "int":
            cell.number_format = '0'

    workbook.save(excel_path)


def generate_prompt(prompt_length):
    vocab = list(tokenizer.get_vocab().keys())
    random.seed(42)
    selected_tokens = random.sample(vocab, prompt_length)
    return tokenizer.decode(tokenizer.convert_tokens_to_ids(selected_tokens))


def send_request(session_id, workbook, sheet, excel_path, prompt):
    metrics = {
        'session_id': session_id,
        'BT': time.time(),
        'FT': None,
        'LT': None,
        'prompt_tokens': None,
        'prompt_ms': None,
        'prompt_per_token_ms': None,
        'prompt_per_second': None,
        'predicted_ms': None,
        'predicted_per_token_ms': None,
        'predicted_per_second': None,
        'content': "",
        'error': None
    }

    try:
        stream = client.chat.completions.create(
            model="deepseek-ai/DeepSeek-R1-Distill-Qwen-1.5B",
            messages=[{"role": "system", "content": prompt}],
            stream=True,
            max_tokens=MAX_TOKENS,
            stream_options={"include_usage": True}
        )

        for chunk in stream:
            if metrics['FT'] is None and chunk.choices[0].delta.content:
                metrics['FT'] = time.time()

            if chunk.choices[0].delta.content:
                metrics['content'] += chunk.choices[0].delta.content
            else:
                if hasattr(chunk, 'usage') and hasattr(chunk, 'model_extra'):
                    usage_data = chunk.usage
                    timings_data = chunk.model_extra.get('timings', {}) if hasattr(chunk, 'model_extra') else {}

                    metrics.update({
                        'prompt_tokens': usage_data.prompt_tokens if hasattr(usage_data, 'prompt_tokens') else None,
                        'prompt_ms': timings_data.get('prompt_ms'),
                        'prompt_per_token_ms': timings_data.get('prompt_per_token_ms'),
                        'prompt_per_second': timings_data.get('prompt_per_second'),
                        'predicted_ms': timings_data.get('predicted_ms'),
                        'predicted_per_token_ms': timings_data.get('predicted_per_token_ms'),
                        'predicted_per_second': timings_data.get('predicted_per_second')
                    })

            if chunk.choices[0].finish_reason:
                metrics['LT'] = time.time()
                break

        if metrics['FT'] and metrics['BT']:
            metrics['TTFT'] = (metrics['FT'] - metrics['BT']) * 1000
        if metrics['LT'] and metrics['FT']:
            metrics['TGT'] = (metrics['LT'] - metrics['FT']) * 1000
        if metrics.get('TTFT') and metrics['prompt_tokens']:
            metrics['PP'] = (metrics['TTFT'] / metrics['prompt_tokens'])

    except Exception as e:
        metrics['error'] = str(e)

    metrics['content_sample'] = (metrics['content'][:100] ) if metrics['content'] else None
    save_to_excel(metrics, workbook, sheet, excel_path)
    return metrics


def run_benchmark_for_config(client_count, prompt_length):
    print(f"\nStarting benchmark with {client_count} clients and {prompt_length} token prompt...")

    prompt = generate_prompt(prompt_length)

    workbook, sheet, excel_path = init_excel_file(client_count, prompt_length)

    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=client_count) as executor:
            futures = {executor.submit(send_request, i, workbook, sheet, excel_path, prompt): i
                       for i in range(client_count)}

            for future in concurrent.futures.as_completed(futures):
                session_id = futures[future]
                try:
                    metrics = future.result()
                    print(f"Session {session_id} completed successfully")
                except Exception as e:
                    print(f"Session {session_id} failed with error: {str(e)}")
    finally:
        workbook.close()
        print(f"Results saved to: {excel_path}")


def run_all_benchmarks():
    print("Starting comprehensive benchmark tests...")

    for client_count in CLIENT_COUNTS:
        for prompt_length in PROMPT_LENGTHS:
            run_benchmark_for_config(client_count, prompt_length)

    print("\nAll benchmark tests completed!")


if __name__ == "__main__":
    run_all_benchmarks()