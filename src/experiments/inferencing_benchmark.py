import openai
import openpyxl
import os
from openpyxl import load_workbook

# OpenAI client configuration
client = openai.OpenAI(
    base_url="http://192.168.88.130:8080",
    api_key="sk-no-key-required"
)

# OpenAI chat completion request
stream = client.chat.completions.create(
    model="deepseek-ai/DeepSeek-R1-Distill-Qwen-1.5B",
    messages=[{"role": "system", "content": "give me a 200 word essay"}],
    stream=True,
    max_tokens=200,
    stream_options={"include_usage": True}
)

# Excel file path
excel_file = r"C:\Users\ASUS\Desktop\Pars\HardwareAware\src\experiments\benchmark.xlsx"

# Ensure the directory exists
directory = os.path.dirname(excel_file)
if not os.path.exists(directory):
    os.makedirs(directory)

# Check if the Excel file exists; if not, create it
try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["id", "prompt_tokens", "prompt_ms", "prompt_per_token_ms", "prompt_per_second",
                  "predicted_ms", "predicted_per_token_ms", "predicted_per_second"])

# Iterate through the stream and extract the necessary details
for chunk in stream:
    if chunk.choices[0].delta.content:  # Check if content is present
        print(chunk.choices[0].delta.content, end="", flush=True)
    else:
        if chunk.usage:  # Ensure usage information is available
            data = chunk.usage
            timings = chunk.timings

            # Extract required data fields from timings (using dictionary keys instead of attributes)
            row = [
                chunk.id,  # ID
                data['prompt_tokens'] if 'prompt_tokens' in data else None,
                timings['prompt_ms'] if 'prompt_ms' in timings else None,
                timings['prompt_per_token_ms'] if 'prompt_per_token_ms' in timings else None,
                timings['prompt_per_second'] if 'prompt_per_second' in timings else None,
                timings['predicted_ms'] if 'predicted_ms' in timings else None,
                timings['predicted_per_token_ms'] if 'predicted_per_token_ms' in timings else None,
                timings['predicted_per_second'] if 'predicted_per_second' in timings else None
            ]

            # Append the extracted data to the Excel sheet
            sheet.append(row)

            # Save the workbook after appending
            workbook.save(excel_file)
