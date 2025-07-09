import openai
import openpyxl
import os
import concurrent.futures
from openpyxl import load_workbook


client = openai.OpenAI(
    base_url="http://192.168.88.130:8080",
    api_key="sk-no-key-required"
)

excel_file = r"C:\Users\ASUS\Desktop\Pars\HardwareAware\src\experiments\benchmark.xlsx"

directory = os.path.dirname(excel_file)
if not os.path.exists(directory):
    os.makedirs(directory)

try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    print(f"Successfully loaded the workbook from {excel_file}")
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["iteration", "id", "prompt_tokens", "prompt_ms", "prompt_per_token_ms", "prompt_per_second",
                  "predicted_ms", "predicted_per_token_ms", "predicted_per_second"])
    print(f"Created a new workbook at {excel_file}")

def run_openai_request(iteration):
    print(f"Running iteration {iteration}...")

    stream = client.chat.completions.create(
        model="deepseek-ai/DeepSeek-R1-Distill-Qwen-1.5B",
        messages=[{"role": "system", "content": "give me a 200 word essay"}],
        stream=True,
        max_tokens=200,
        stream_options={"include_usage": True}
    )

    for chunk in stream:
        if chunk.choices[0].delta.content:
            print(chunk.choices[0].delta.content, end="", flush=True)
        else:
            if 'usage' in chunk and 'timings' in chunk:
                usage_data = chunk['usage']
                timings_data = chunk['timings']

                prompt_tokens = usage_data.get('prompt_tokens', None)
                prompt_ms = timings_data.get('prompt_ms', None)
                prompt_per_token_ms = timings_data.get('prompt_per_token_ms', None)
                prompt_per_second = timings_data.get('prompt_per_second', None)
                predicted_ms = timings_data.get('predicted_ms', None)
                predicted_per_token_ms = timings_data.get('predicted_per_token_ms', None)
                predicted_per_second = timings_data.get('predicted_per_second', None)

                row = [
                    iteration,
                    chunk['id'],
                    prompt_tokens,
                    prompt_ms,
                    prompt_per_token_ms,
                    prompt_per_second,
                    predicted_ms,
                    predicted_per_token_ms,
                    predicted_per_second
                ]

                print(f"Appending row: {row}")

                sheet.append(row)

                workbook.save(excel_file)
                print(f"Workbook saved after iteration {iteration}.")

    print(f"Iteration {iteration} completed.\n")

with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
    futures = [executor.submit(run_openai_request, iteration) for iteration in range(1, 2)]

    concurrent.futures.wait(futures)

print("Process completed with double requests at the same time.")
